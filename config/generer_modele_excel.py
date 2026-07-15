"""
Génère le fichier modèle pour la surcharge externe (config/surcharge_config.py).
==================================================================================
À lancer une fois pour produire docs/modele_config_prospection.xlsx, pré-rempli
avec les valeurs actuelles de config/cibles.py.

Ce .xlsx n'est PAS destiné à OneDrive (le tenant Microsoft 365 de G2S interdit
les liens de partage anonymes) : Pauline doit l'IMPORTER dans un nouveau Google
Sheet (Fichier > Importer > Charger), l'ajuster, puis le PARTAGER (bouton
Partager > Général > "Toute personne disposant du lien" > rôle Lecteur) pour
obtenir le lien à mettre dans CONFIG_SHEET_URL (.env) — n'importe quel lien
Google Sheets standard convient, converti automatiquement en export CSV.

Usage (depuis la racine du projet) :
    python config/generer_modele_excel.py

Ne fait PAS partie de l'exécution normale de l'agent (main.py) : c'est un
utilitaire à lancer à la main, une fois, ou de nouveau si les valeurs par
défaut de config/cibles.py changent et que le modèle doit être régénéré.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

RACINE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CHEMIN_SORTIE = os.path.join(RACINE, "docs", "modele_config_prospection.xlsx")

# (variable, valeur par défaut actuelle de cibles.py, description pour Pauline)
LIGNES = [
    ("section_activite_principale", "E,F,H,I,N,Q",
     "Sections NAF ciblées (E=déchets, F=BTP, H=transport, I=HCR, N=intérim/propreté/sécurité, Q=santé)."),
    ("tranche_effectif_salarie", "21,22,31,32,41,42,51,52,53",
     "Codes INSEE de tranche d'effectif ciblés (50 salariés et plus)."),
    ("categorie_entreprise", "PME,ETI,GE",
     "Catégories d'entreprise ciblées."),
    ("etat_administratif", "A",
     "A = entreprises actives uniquement."),
    ("departements_cible", "75,92,93,94,77,78,91,95,971,972,973,974,975,976",
     "Départements parcourus tour à tour (un par exécution), séparés par des virgules."),
    ("objectif_par_run", 50,
     "Nombre de nouveaux prospects visés par exécution."),
    ("seuil_retention", "",
     "Score minimum pour garder un prospect. Laisser VIDE = garder tous les prospects."),
    ("injecter_pipedrive", "Oui",
     "Oui/Non — injecter les prospects retenus dans Pipedrive."),
    ("exclure_procedure_collective", "Oui",
     "Oui/Non — exclure les entreprises en procédure collective (BODACC)."),
    ("notifier", "Oui",
     "Oui/Non — envoyer le récapitulatif de fin d'exécution."),
]


def generer(chemin=CHEMIN_SORTIE):
    os.makedirs(os.path.dirname(chemin), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Config"

    entete_fill = PatternFill("solid", fgColor="1F4E78")
    entete_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    data_font = Font(name="Arial", size=10)

    for c, titre in enumerate(("Variable", "Valeur", "Description"), start=1):
        cell = ws.cell(row=1, column=c, value=titre)
        cell.fill, cell.font = entete_fill, entete_font
        cell.alignment = Alignment(vertical="center")

    for r, (variable, valeur, description) in enumerate(LIGNES, start=2):
        ws.cell(row=r, column=1, value=variable).font = data_font
        ws.cell(row=r, column=2, value=valeur).font = data_font
        cell_desc = ws.cell(row=r, column=3, value=description)
        cell_desc.font = data_font

    ws.cell(row=1, column=1).comment = Comment(
        "Ne pas renommer les variables de la colonne A : ce sont les noms attendus par "
        "config/surcharge_config.py. Laisser une cellule Valeur vide = garder le réglage "
        "actuel du code, ne rien changer. Une fois importé dans Google Sheets : bouton "
        "Partager > Général > \"Toute personne disposant du lien\" > rôle Lecteur, puis "
        "copier le lien.", "Agent prospection")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    largeurs = {1: 32, 2: 45, 3: 75}
    for c, largeur in largeurs.items():
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = largeur

    wb.save(chemin)
    print(f"Modèle généré -> {chemin}")


if __name__ == "__main__":
    generer()
