"""
Surcharge des paramètres de config/cibles.py via une source externe.
======================================================================
But : permettre à Pauline de modifier le ciblage sans toucher au code.

Source de PRODUCTION (conteneur Docker chez un hébergeur, sans accès au disque
de Pauline) : un GOOGLE SHEET. Choisi après avoir constaté que le tenant
Microsoft 365 de G2S interdit les liens de partage OneDrive/SharePoint
anonymes ("Toute personne disposant du lien") — un Google Sheet partagé de la
même façon n'est pas concerné par cette politique (compte Google, pas le
tenant G2S). CONFIG_SHEET_URL (.env) accepte N'IMPORTE QUEL lien Google Sheets
standard (le lien "Copier le lien" du bouton Partager, un lien de publication,
etc.) : il est automatiquement converti vers l'URL d'export CSV. La seule
condition est que le fichier soit partagé en lecture pour "Toute personne
disposant du lien" (bouton Partager > Général > "Toute personne disposant du
lien" > rôle Lecteur) — pas "Accès restreint", sinon Google redirige vers une
page de connexion que le conteneur ne peut pas franchir.

Chemin LOCAL (fichier .xlsx sur disque) toujours accepté en plus, uniquement
pour des tests lancés hors conteneur.

Format attendu (Google Sheet : 1re feuille publiée ; fichier local : 1re
feuille du classeur) :
    Ligne 1  : en-têtes "Variable" | "Valeur" (ignorée si présente)
    Lignes suivantes : nom de variable | valeur

Variables reconnues (voir config/generer_modele_excel.py pour un modèle prêt à
importer dans Google Sheets) :
    section_activite_principale, tranche_effectif_salarie, categorie_entreprise,
    etat_administratif        -> clés de FILTRES_CIBLE (texte, ex. "E,F,H,I,N,Q")
    departements_cible        -> DEPARTEMENTS_CIBLE (codes séparés par des virgules)
    objectif_par_run          -> OBJECTIF_PAR_RUN (nombre entier)
    seuil_retention            -> SEUIL_RETENTION (nombre entier, vide = pas de seuil)
    injecter_pipedrive, exclure_procedure_collective, notifier
                               -> booléens (écrire "Oui" ou "Non")

Règle de repli : une variable ABSENTE de la source, ou dont la cellule est
VIDE, laisse la valeur de config/cibles.py inchangée. Une valeur présente mais
du mauvais type (ex. du texte à la place d'un nombre) est ignorée avec un
message -> la source externe ne peut jamais faire planter l'agent, au pire
elle est sans effet.
"""

import os
import io
import csv
import re
import unicodedata
import requests
from openpyxl import load_workbook

NOM_VARIABLE_ENV = "CONFIG_SHEET_URL"
_RE_ID_GOOGLE_SHEET = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")

# Clés de FILTRES_CIBLE modifiables depuis la source externe (texte brut, même
# convention que dans cibles.py : valeurs multiples séparées par des virgules).
CLES_FILTRES = ["section_activite_principale", "tranche_effectif_salarie",
               "categorie_entreprise", "etat_administratif"]


def _sans_accents(texte):
    return "".join(c for c in unicodedata.normalize("NFKD", texte) if not unicodedata.combining(c))


def _source_externe(racine):
    """URL du Google Sheet publié (ou chemin local pour les tests) : variable
    d'environnement, sinon .env."""
    valeur = os.environ.get(NOM_VARIABLE_ENV)
    if valeur:
        return valeur.strip()
    chemin_env = os.path.join(racine, ".env")
    if os.path.exists(chemin_env):
        with open(chemin_env, encoding="utf-8") as f:
            for ligne in f:
                ligne = ligne.strip()
                if ligne.startswith(NOM_VARIABLE_ENV) and "=" in ligne:
                    valeur = ligne.split("=", 1)[1].strip().strip('"').strip("'")
                    return valeur or None
    return None


def _est_url(source):
    return str(source).lower().startswith(("http://", "https://"))


def _lire_lignes_depuis_paires(paires):
    """`paires` : itérable de (cle_brute, valeur_brute) -> dict filtré/normalisé,
    en ignorant l'en-tête et les lignes sans valeur."""
    lignes = {}
    for cle_brute, valeur in paires:
        if cle_brute is None:
            continue
        cle = str(cle_brute).strip()
        if cle.lower() in ("variable", ""):
            continue  # en-tête ou ligne vide
        if valeur is None or str(valeur).strip() == "":
            continue  # cellule vide -> pas de surcharge pour cette ligne
        lignes[_sans_accents(cle).lower()] = str(valeur).strip()
    return lignes


def _url_export_csv(url):
    """Convertit n'importe quel lien Google Sheets standard (édition, partage,
    publication) vers l'URL d'export CSV de la 1re feuille. Un lien d'export ou
    de publication déjà construit est laissé tel quel."""
    if "output=csv" in url or "/export" in url:
        return url
    correspondance = _RE_ID_GOOGLE_SHEET.search(url)
    if not correspondance:
        return url  # pas reconnu comme un lien Google Sheets standard
    id_feuille = correspondance.group(1)
    gid = re.search(r"gid=(\d+)", url)
    return f"https://docs.google.com/spreadsheets/d/{id_feuille}/export?format=csv&gid={gid.group(1) if gid else '0'}"


def _lire_lignes_csv(url):
    """Télécharge et parse un Google Sheet en CSV. Nécessite que le fichier soit
    partagé en lecture pour "Toute personne disposant du lien" — sinon Google
    répond par une page de connexion (HTML), pas le contenu du tableau."""
    reponse = requests.get(_url_export_csv(url), timeout=30)
    reponse.raise_for_status()
    if "text/csv" not in (reponse.headers.get("Content-Type") or ""):
        raise ValueError("réponse reçue non-CSV — le fichier est-il bien partagé en "
                        "lecture pour \"Toute personne disposant du lien\" ?")
    texte = reponse.content.decode("utf-8-sig")
    paires = ((row[0] if row else None, row[1] if len(row) > 1 else None)
             for row in csv.reader(io.StringIO(texte)))
    return _lire_lignes_depuis_paires(paires)


def _lire_lignes_excel(chemin):
    """Fichier .xlsx local — uniquement pour tester hors conteneur."""
    wb = load_workbook(chemin, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    paires = ((row[0] if row else None, row[1] if row and len(row) > 1 else None)
             for row in ws.iter_rows(values_only=True))
    return _lire_lignes_depuis_paires(paires)


def _vers_entier(valeur, nom):
    try:
        return int(str(valeur).strip())
    except (TypeError, ValueError):
        print(f"  [!] Config externe : « {nom} » = {valeur!r} n'est pas un nombre entier, ignoré.")
        return None


def _vers_booleen(valeur, nom):
    texte = _sans_accents(str(valeur).strip()).lower()
    if texte in ("oui", "vrai", "true", "1"):
        return True
    if texte in ("non", "faux", "false", "0"):
        return False
    print(f"  [!] Config externe : « {nom} » = {valeur!r} attendu \"Oui\" ou \"Non\", ignoré.")
    return None


def appliquer_surcharges(valeurs, racine):
    """`valeurs` : dict des constantes par défaut de cibles.py (FILTRES_CIBLE,
    DEPARTEMENTS_CIBLE, OBJECTIF_PAR_RUN, SEUIL_RETENTION, INJECTER_PIPEDRIVE,
    EXCLURE_PROCEDURE_COLLECTIVE, NOTIFIER). Renvoie un dict de même forme, avec
    les valeurs de la source externe appliquées quand elles sont définies et
    valides (sinon les valeurs par défaut sont conservées telles quelles), plus
    une clé "_source" (texte lisible décrivant d'où viennent les valeurs —
    repris dans le récapitulatif de fin d'exécution, voir SOURCE_CONFIG dans
    cibles.py)."""
    source = _source_externe(racine)
    if not source:
        valeurs["_source"] = "config/cibles.py (aucune source externe configurée)"
        return valeurs
    if not _est_url(source) and not os.path.exists(source):
        print(f"  [!] {NOM_VARIABLE_ENV} pointe vers un fichier local introuvable ({source}) — "
              f"config/cibles.py utilisé tel quel.")
        valeurs["_source"] = f"config/cibles.py (fichier local introuvable : {source})"
        return valeurs

    try:
        lignes = _lire_lignes_csv(source) if _est_url(source) else _lire_lignes_excel(source)
    except Exception as e:
        print(f"  [!] Config externe illisible ({e}) — config/cibles.py utilisé tel quel. "
              f"Si {NOM_VARIABLE_ENV} est un Google Sheet, vérifier qu'il est bien partagé en "
              f"lecture pour \"Toute personne disposant du lien\" (bouton Partager > Général).")
        valeurs["_source"] = f"config/cibles.py (source externe illisible : {source})"
        return valeurs

    if not lignes:
        print(f"  [!] Config externe : 0 variable reconnue dans {source} — config/cibles.py "
              f"utilisé tel quel. Vérifier le partage (\"Toute personne disposant du lien\") "
              f"et les noms de variables en colonne A.")
        valeurs["_source"] = f"config/cibles.py (0 variable reconnue dans {source})"
        return valeurs

    resultat = dict(valeurs)

    filtres = dict(resultat["FILTRES_CIBLE"])
    for cle in CLES_FILTRES:
        if cle in lignes:
            filtres[cle] = lignes[cle]
    resultat["FILTRES_CIBLE"] = filtres

    if "departements_cible" in lignes:
        depts = [d.strip() for d in lignes["departements_cible"].split(",") if d.strip()]
        if depts:
            resultat["DEPARTEMENTS_CIBLE"] = depts

    if "objectif_par_run" in lignes:
        entier = _vers_entier(lignes["objectif_par_run"], "objectif_par_run")
        if entier is not None:
            resultat["OBJECTIF_PAR_RUN"] = entier

    if "seuil_retention" in lignes:
        entier = _vers_entier(lignes["seuil_retention"], "seuil_retention")
        if entier is not None:
            resultat["SEUIL_RETENTION"] = entier

    for cle, nom_constante in (("injecter_pipedrive", "INJECTER_PIPEDRIVE"),
                              ("exclure_procedure_collective", "EXCLURE_PROCEDURE_COLLECTIVE"),
                              ("notifier", "NOTIFIER")):
        if cle in lignes:
            booleen = _vers_booleen(lignes[cle], cle)
            if booleen is not None:
                resultat[nom_constante] = booleen

    print(f"  [i] Config externe appliquée depuis {source}.")
    origine = "Google Sheet" if _est_url(source) else "fichier local"
    resultat["_source"] = f"{origine} ({source})"
    return resultat
