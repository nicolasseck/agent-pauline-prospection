"""
Export Excel — sortie commune de l'agent (fichier lisible et filtrable).
N'affiche que les colonnes déclarées dans ENTETES (ordre = ordre de ENTETES).
Les colonnes "lien_*" sont rendues cliquables (fonctionne dans Excel ET, après
import, dans Google Sheets).
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.referentiels import ENTETES

# Colonnes de type lien -> libellé cliquable affiché dans la cellule
LIBELLES_LIENS = {
    "lien_linkedin": "Ouvrir LinkedIn",
    "lien_google": "Ouvrir Google",
    "lien_site_web": "Ouvrir le site internet",
}

# Colonnes de type lien affichant l'URL complète (pas de libellé) : la cellule
# reste cliquable, mais montre le https://... tel quel.
COLONNES_LIEN_URL_COMPLETE = {"conventions_idcc_url"}


def exporter_xlsx(fiches, chemin):
    """Écrit une liste de fiches (dicts) dans un fichier Excel mis en forme."""
    if not fiches:
        print("Aucune fiche à exporter.")
        return
    os.makedirs(os.path.dirname(chemin), exist_ok=True)

    colonnes = [cle for cle in ENTETES if any(cle in f for f in fiches)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"

    entete_fill = PatternFill("solid", fgColor="1F4E78")
    entete_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    data_font = Font(name="Arial", size=10)
    lien_font = Font(name="Arial", size=10, color="0563C1", underline="single")

    for c, cle in enumerate(colonnes, start=1):
        cell = ws.cell(row=1, column=c, value=ENTETES[cle])
        cell.fill, cell.font = entete_fill, entete_font
        cell.alignment = Alignment(vertical="center")

    for r, fiche in enumerate(fiches, start=2):
        for c, cle in enumerate(colonnes, start=1):
            cell = ws.cell(row=r, column=c)
            valeur = fiche.get(cle)
            if cle in LIBELLES_LIENS and valeur:
                cell.value = LIBELLES_LIENS[cle]
                cell.hyperlink = valeur
                cell.font = lien_font
            elif cle in COLONNES_LIEN_URL_COMPLETE and valeur and "; " not in str(valeur):
                cell.value = valeur  # un seul lien -> cellule cliquable
                cell.hyperlink = valeur
                cell.font = lien_font
            else:
                cell.value = valeur
                cell.font = data_font
                if cle == "chiffre_affaires" and valeur is not None:
                    cell.number_format = "#,##0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for c, cle in enumerate(colonnes, start=1):
        if cle in LIBELLES_LIENS:
            largeur = len(LIBELLES_LIENS[cle])
        else:
            largeur = max([len(str(ENTETES[cle]))] +
                          [len(str(f.get(cle) or "")) for f in fiches])
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(largeur + 2, 55)

    wb.save(chemin)
    print(f"{len(fiches)} fiche(s) exportée(s) -> {chemin}")
