"""
Référentiels — tables de correspondance (codes INSEE -> libellés lisibles).
Données de référence stables ; rarement modifiées (sauf les conventions
collectives, chargées depuis data/convetions_c/convention.json).
"""

import os
import json

# --- Tranches d'effectif salarié (codes INSEE) ----------------------------------
# Cible G2S "50 à 500 salariés" => codes ["21", "22", "31", "32"].
TRANCHES_EFFECTIF = {
    "NN": "Non renseigné / non employeur", "00": "0 salarié",
    "01": "1 à 2 salariés", "02": "3 à 5 salariés", "03": "6 à 9 salariés",
    "11": "10 à 19 salariés", "12": "20 à 49 salariés", "21": "50 à 99 salariés",
    "22": "100 à 199 salariés", "31": "200 à 249 salariés", "32": "250 à 499 salariés",
    "41": "500 à 999 salariés", "42": "1 000 à 1 999 salariés",
    "51": "2 000 à 4 999 salariés", "52": "5 000 à 9 999 salariés",
    "53": "10 000 salariés et plus",
}

# --- Sections d'activité (NAF niveau 1) -----------------------------------------
SECTIONS = {
    "A": "Agriculture, sylviculture, pêche", "B": "Industries extractives",
    "C": "Industrie manufacturière", "D": "Énergie", "E": "Eau, déchets",
    "F": "Construction (BTP)", "G": "Commerce, réparation auto",
    "H": "Transports et entreposage", "I": "Hébergement et restauration",
    "J": "Information et communication", "K": "Activités financières et d'assurance",
    "L": "Activités immobilières", "M": "Activités scientifiques et techniques",
    "N": "Services administratifs et de soutien", "O": "Administration publique",
    "P": "Enseignement", "Q": "Santé humaine et action sociale",
    "R": "Arts, spectacles, loisirs", "S": "Autres activités de services",
    "T": "Activités des ménages employeurs", "U": "Activités extra-territoriales",
}

# --- Conventions collectives (codes IDCC) : nom usuel + lien Légifrance --------
# Source : jeu de données data/convetions_c/convention.json (export Légifrance,
# une ligne par texte : convention de base, avenant ou accord). Chargé une seule
# fois puis mis en cache. Pour chaque IDCC, on retient le texte "CONVENTION
# COLLECTIVE..." toujours EN VIGUEUR s'il existe (pas un avenant ni un texte
# abrogé/dénoncé/périmé), sinon le meilleur texte disponible pour ce code.
# Un code absent du fichier est affiché tel quel, sans nom ni lien inventés.
_CHEMIN_CONVENTIONS = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                   "data", "convetions_c", "convention.json")
_CONVENTIONS_IDCC_CACHE = None


def _rang_texte(nature, etat):
    """Plus le rang est élevé, plus le texte est un bon candidat pour représenter
    la convention : texte de convention (pas un avenant/accord isolé) ET en
    vigueur (pas abrogé/dénoncé/périmé/remplacé)."""
    nature = (nature or "").upper()
    etat = (etat or "").upper()
    return ("CONVENTION COLLECTIVE" in nature, etat.startswith("VIGUEUR"))


def _charger_conventions_idcc():
    global _CONVENTIONS_IDCC_CACHE
    if _CONVENTIONS_IDCC_CACHE is not None:
        return _CONVENTIONS_IDCC_CACHE
    meilleurs = {}
    try:
        with open(_CHEMIN_CONVENTIONS, encoding="utf-8") as f:
            lignes = json.load(f)
        for ligne in lignes:
            code, titre, url = ligne.get("Unnamed: 3"), ligne.get("Unnamed: 4"), ligne.get("Unnamed: 9")
            if not code or not titre or not url:
                continue
            code = str(code).strip().zfill(4)
            rang = _rang_texte(ligne.get("Unnamed: 5"), ligne.get("Unnamed: 6"))
            if code not in meilleurs or rang > meilleurs[code][0]:
                meilleurs[code] = (rang, titre, url)
    except (OSError, json.JSONDecodeError, ValueError) as e:
        print(f"  [!] {_CHEMIN_CONVENTIONS} illisible ({e}) — noms/liens de CCN indisponibles.")
    _CONVENTIONS_IDCC_CACHE = {code: (titre, url) for code, (_, titre, url) in meilleurs.items()}
    return _CONVENTIONS_IDCC_CACHE


# Compléments manuels : codes IDCC importants pour les secteurs prioritaires de
# Pauline mais ABSENTS de data/convetions_c/convention.json — pas parce que le
# fichier est incomplet dans son ensemble, mais parce que le texte fondateur de
# ces conventions y est classé "TI" (texte indépendant) SANS numéro IDCC sur sa
# ligne (ex. IDCC 1413/intérim : founding accord de 1986, présent dans le
# fichier mais avec un champ IDCC vide -> invisible pour le chargeur ci-dessus).
# Chaque entrée vérifiée individuellement (identifiant KALICONT Légifrance).
# Prioritaire sur le fichier JSON. Table volontairement non exhaustive, à
# compléter à chaque nouveau trou constaté (voir README, journal d'avancement).
_CONVENTIONS_IDCC_MANUEL = {
    "1413": ("Travail temporaire (salariés permanents)",
             "https://www.legifrance.gouv.fr/conv_coll/id/KALICONT000005635905"),
}


def libelle_idcc(code):
    """Nom usuel de la convention collective pour un code IDCC, si connu."""
    code_norm = str(code).strip().zfill(4)
    if code_norm in _CONVENTIONS_IDCC_MANUEL:
        return _CONVENTIONS_IDCC_MANUEL[code_norm][0]
    infos = _charger_conventions_idcc().get(code_norm)
    return infos[0] if infos else None


def lien_legifrance_idcc(code):
    """URL Légifrance de la convention collective pour un code IDCC, si connue."""
    code_norm = str(code).strip().zfill(4)
    if code_norm in _CONVENTIONS_IDCC_MANUEL:
        return _CONVENTIONS_IDCC_MANUEL[code_norm][1]
    infos = _charger_conventions_idcc().get(code_norm)
    return infos[1] if infos else None


# --- Suggestion de CCN par secteur (DARES) — UNIQUEMENT quand l'entreprise -----
# n'a AUCUNE convention officiellement déclarée par l'INSEE. Source :
# data/dares_ape_idcc.xlsx (table de passage DARES code APE -> IDCC, données
# 2022) : pour chaque code APE, plusieurs conventions coexistent en pratique,
# chacune avec un % de salariés du secteur couverts (somme = 100 % par code
# APE). On retient la plus fréquente et on affiche CE POURCENTAGE, pour que ce
# soit toujours lu comme une indication statistique, jamais une certitude.
# Certains codes APE sont "non diffusables" (échantillon trop petit / faible
# couverture par les IDCC de branche) : pas de suggestion dans ce cas, jamais
# de valeur inventée.
_CHEMIN_DARES_APE_IDCC = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                      "data", "dares_ape_idcc.xlsx")
_FEUILLE_DARES = "IDCC2022_passageAPEIDCC_diff"
_SUGGESTIONS_APE_IDCC_CACHE = None


def _charger_suggestions_ape_idcc():
    global _SUGGESTIONS_APE_IDCC_CACHE
    if _SUGGESTIONS_APE_IDCC_CACHE is not None:
        return _SUGGESTIONS_APE_IDCC_CACHE
    meilleurs = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(_CHEMIN_DARES_APE_IDCC, read_only=True, data_only=True)
        lignes = wb[_FEUILLE_DARES].iter_rows(values_only=True)
        for ligne in lignes:
            if ligne and ligne[0] == "apen":
                break  # fin des lignes d'intro/mode d'emploi, données juste après
        for ligne in lignes:
            if not ligne or len(ligne) < 6:
                continue
            apen, _intape, _eff, idcc, intidcc, pctdiff = ligne[:6]
            if not apen or not idcc or idcc == "Autre" or not isinstance(pctdiff, (int, float)):
                continue  # code non diffusable, regroupement "Autre", ou ligne mal formée
            code_naf = "".join(ch for ch in str(apen) if ch.isalnum()).upper()
            try:
                code_idcc = str(int(idcc)).zfill(4)
            except (TypeError, ValueError):
                continue
            existant = meilleurs.get(code_naf)
            if existant is None or pctdiff > existant[2]:
                meilleurs[code_naf] = (code_idcc, intidcc, pctdiff)
    except (OSError, KeyError, ValueError) as e:
        print(f"  [!] {_CHEMIN_DARES_APE_IDCC} illisible ({e}) — suggestions de CCN par secteur indisponibles.")
    _SUGGESTIONS_APE_IDCC_CACHE = meilleurs
    return _SUGGESTIONS_APE_IDCC_CACHE


def suggestion_idcc_pour_naf(code_naf):
    """(code_idcc, intitulé DARES, % de salariés du secteur couverts) pour ce
    code NAF/APE, si connu de la table DARES — sinon None (code non diffusable
    ou absent de la table). À utiliser UNIQUEMENT en absence de CCN officielle
    (voir collecte.py) : c'est une indication statistique, pas une certitude."""
    if not code_naf:
        return None
    code = "".join(ch for ch in str(code_naf) if ch.isalnum()).upper()
    return _charger_suggestions_ape_idcc().get(code)


# --- Secteurs prioritaires (grille Pauline) -------------------------------------
# Préfixe de code NAF (sans point) -> (libellé, poids). Poids : Très élevée=4, Élevée=3.
SECTEURS_PRIORITAIRES = [
    ("38", "Déchets", 4), ("39", "Déchets", 4),
    ("812", "Propreté", 4),
    ("782", "Intérim", 4), ("783", "Intérim", 4),
    ("49", "Transport-logistique", 4), ("52", "Transport-logistique", 4), ("53", "Transport-logistique", 4),
    ("41", "BTP", 4), ("42", "BTP", 4), ("43", "BTP", 4),
    ("86", "Santé & médico-social", 4), ("87", "Santé & médico-social", 4), ("88", "Santé & médico-social", 4),
    ("80", "Sécurité privée", 4),
    ("55", "HCR", 4), ("56", "HCR", 4),
    ("822", "Services aux entreprises", 3),
]
# Secteurs à convention collective complexe (sous-ensemble, défini par Pauline)
SECTEURS_COMPLEXES = {"Transport-logistique", "BTP", "HCR", "Déchets", "Intérim"}


def secteur_prioritaire(naf):
    """Renvoie (libellé, poids) si le code NAF appartient à un secteur prioritaire."""
    if not naf:
        return None
    code = "".join(ch for ch in naf if ch.isalnum()).upper()
    for prefixe, label, poids in sorted(SECTEURS_PRIORITAIRES, key=lambda x: -len(x[0])):
        if code.startswith(prefixe):
            return (label, poids)
    return None


# --- Formes juridiques (catégorie juridique INSEE niveau III, principales) -------
FORMES_JURIDIQUES = {
    "1000": "Entrepreneur individuel",
    "5202": "Société en nom collectif (SNC)",
    "5306": "Société en commandite simple",
    "5499": "SARL", "5498": "SARL unipersonnelle (EURL)",
    "5410": "SARL nationale", "5458": "SCOP (SARL)",
    "5505": "SA à participation ouvrière (CA)", "5510": "SA à conseil d'administration",
    "5515": "SA d'économie mixte (CA)", "5560": "SA à directoire", "5599": "SA",
    "5710": "SAS", "5720": "SAS à associé unique (SASU)",
    "5785": "Société d'exercice libéral par actions simplifiée (SELAS)",
    "6540": "SCI", "6220": "GIE",
    "9210": "Association non déclarée", "9220": "Association déclarée",
    "9221": "Association déclarée (insertion par l'économique)",
    "9230": "Association déclarée reconnue d'utilité publique",
    "9240": "Congrégation", "9260": "Association de droit local (Alsace-Moselle)",
    "7344": "Établissement public local",
}

# --- En-têtes Excel, DANS L'ORDRE D'AFFICHAGE -----------------------------------
# L'export n'affiche que les clés présentes ici (les champs techniques internes,
# ex. "section" ou "tranche_effectif_code", sont donc automatiquement masqués).
ENTETES = {
    "score": "Score (priorité)", "score_raison": "Critères validés",
    "siren": "SIREN", "raison_sociale": "Raison sociale", "naf": "Code NAF",
    "forme_juridique": "Forme juridique", "categorie": "Catégorie",
    "tranche_effectif": "Effectif (tranche)", "chiffre_affaires": "Chiffre d'affaires (€)",
    "conventions_idcc": "Convention(s) collective(s)", "conventions_idcc_url": "Lien Légifrance (CCN)",
    "conv_renseignee": "Conv. renseignée",
    "etat": "État", "adresse": "Adresse", "code_postal": "Code postal",
    "ville": "Ville", "departement": "Dpt", "dirigeants": "Dirigeants (mandataires sociaux)",
    "lien_site_web": "Site internet de l'entreprise",
    "lien_linkedin": "Recherche LinkedIn (DRH/DAF)", "lien_google": "Recherche Google (DRH/DAF)",
}
